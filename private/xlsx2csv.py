import pandas as pd
import os
from openpyxl import load_workbook

def xlsx2csv(xlsx_path, drop_columns, output_dir=None):
    # If no output directory is specified, use the same directory as the .xlsx file
    if output_dir is None:
        output_dir = os.path.dirname(xlsx_path)
    os.makedirs(output_dir, exist_ok=True)

    # Load the entire Excel file (all sheets)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    for sheet_name in sheet_names:
        try:
            ws = wb[sheet_name]
            data = list(ws.values)
        except TypeError as e:
            if "DataValidation.__init__" in str(e):
                print(f"Skipped sheet {sheet_name}: DataValidation error ({e})")
                continue
            else:
                raise e

        if not data:
            print(f"Skipped empty sheet: {sheet_name}")
            continue

        # 第一行作为表头
        df = pd.DataFrame(data[1:], columns=data[0])

        # 删除整行为空的行
        df.dropna(how="all", inplace=True)

        # Skip empty sheets
        if df.empty:
            print(f"Skipped empty sheet: {sheet_name}")
            continue

        # Delete the specified columns
        if drop_columns:
            # Only delete the columns that exist in the DataFrame (to avoid KeyError)
            to_drop = [col for col in drop_columns if col in df.columns]
            if to_drop:
                df.drop(columns=to_drop, inplace=True)

        # Create a safe filename (replace invalid characters)
        safe_name = "".join(c if c.isalnum() or c in (' ', '_', '-') else "_" for c in sheet_name) # type: ignore
        csv_path = os.path.join(output_dir, f"{safe_name}.csv")

        # Save the DataFrame as a CSV file (UTF-8 encoding with BOM)
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        print(f"Saved: {csv_path}")

if __name__ == "__main__":
    FILE_NAME = "literature_pool.xlsx"
    FILE_DIR = ["doc", "annotated_data"]
    DROP_COLUMNS = ["PaperID", "BibTeX", "RelatedWork?", "Remark", "RootPaperID", "SnowballingMethod"]

    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(root_dir, *FILE_DIR, FILE_NAME)

    xlsx2csv(file_path, DROP_COLUMNS)
